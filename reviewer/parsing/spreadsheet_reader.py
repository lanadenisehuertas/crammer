import csv
import io

from openpyxl import load_workbook

from reviewer.parsing.base import ParsedContent

_HEADER_TERMS = {"term", "terms", "word", "front", "question"}
_HEADER_DEFS = {"definition", "definitions", "meaning", "back", "answer"}


def _looks_numeric(value: str) -> bool:
    """True if the value is a bare number / date-like token (no descriptive words)."""
    stripped = value.replace(",", "").replace("$", "").replace("%", "").strip()
    if not stripped:
        return True
    # Any alphabetic character means it reads as words, not a bare number/date.
    return not any(ch.isalpha() for ch in stripped)


def _rows_to_content(rows: list[list[str]]) -> ParsedContent:
    rows = [[(c or "").strip() for c in row] for row in rows if any((c or "").strip() for c in row)]
    text = "\n".join("\t".join(row) for row in rows)

    pairs: list[tuple[str, str]] = []
    if rows and all(len(row) == 2 for row in rows):
        data_rows = rows
        first = rows[0]
        has_header = (first[0].lower() in _HEADER_TERMS
                      and first[1].lower() in _HEADER_DEFS)
        if has_header:
            data_rows = rows[1:]
        candidate = [(r[0], r[1]) for r in data_rows if r[0] and r[1]]
        # Trust an explicit term/definition header. Otherwise only treat the
        # sheet as flashcards when the definition column reads like words
        # (a majority non-numeric), so data such as "Date,Amount" is not
        # mistaken for term/definition pairs.
        if candidate and (has_header or _mostly_words([d for _, d in candidate])):
            pairs = candidate
    return ParsedContent(text=text, flashcard_pairs=pairs)


def _mostly_words(values: list[str]) -> bool:
    """True when at least half the values contain descriptive (non-numeric) text."""
    wordy = sum(1 for v in values if not _looks_numeric(v))
    return wordy * 2 >= len(values)


def read_csv(data: bytes) -> ParsedContent:
    text = data.decode("utf-8", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    return _rows_to_content(rows)


def read_xlsx(data: bytes) -> ParsedContent:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[list[str]] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if v is None else str(v) for v in row])
    finally:
        wb.close()
    return _rows_to_content(rows)
